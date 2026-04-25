from flask import Flask, request, jsonify, send_file
import base64, io, os, uuid, time, threading
from datetime import datetime
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.drawing.spreadsheet_drawing import OneCellAnchor, AnchorMarker
from openpyxl.drawing.xdr import XDRPositiveSize2D
from openpyxl.utils.units import pixels_to_EMU
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from PIL import Image as PILImage

app = Flask(__name__)

API_KEY  = os.environ.get("API_KEY", "")
TMP_DIR  = "/tmp/excel_files"
os.makedirs(TMP_DIR, exist_ok=True)

IMG_PX = 52
ROW_PT = 42
DARK   = "1F2D4E"

_file_store = {}
_store_lock = threading.Lock()

def _cleanup_old_files(max_age_seconds=600):
    now = time.time()
    with _store_lock:
        expired = [t for t, v in _file_store.items() if now - v["created_at"] > max_age_seconds]
        for token in expired:
            try:
                os.remove(_file_store[token]["path"])
            except Exception:
                pass
            del _file_store[token]

def check_api_key():
    if not API_KEY:
        return False, "Sunucuda API_KEY tanimlanmamis"
    if request.headers.get("X-API-Key") != API_KEY:
        return False, "Yetkisiz erisim"
    return True, None

def header_style(cell):
    cell.font      = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    cell.fill      = PatternFill("solid", start_color=DARK)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    s = Side(style="thin", color="FFFFFF")
    cell.border    = Border(left=s, right=s, bottom=s)

def row_style(cell, row_no):
    cell.fill      = PatternFill("solid", start_color="F5F7FA" if row_no % 2 == 0 else "FFFFFF")
    cell.font      = Font(name="Arial", size=9)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    s = Side(style="thin", color="D0D7E3")
    cell.border    = Border(left=s, right=s, top=s, bottom=s)

def total_style(cell):
    cell.font      = Font(name="Arial", bold=True, size=10, color=DARK)
    cell.fill      = PatternFill("solid", start_color="E8ECF4")
    cell.alignment = Alignment(horizontal="center", vertical="center")
    s = Side(style="medium", color=DARK)
    cell.border    = Border(left=s, right=s, top=s, bottom=s)

def place_image(ws, b64_str, col, row, px):
    raw = base64.b64decode(b64_str)
    pil = PILImage.open(io.BytesIO(raw)).convert("RGB")
    pil.thumbnail((px, px), PILImage.LANCZOS)
    buf = io.BytesIO()
    pil.save(buf, format="JPEG", quality=85)
    buf.seek(0)
    xl = XLImage(buf)
    xl.width  = pil.width
    xl.height = pil.height
    offset = pixels_to_EMU(2)
    size   = XDRPositiveSize2D(pixels_to_EMU(pil.width), pixels_to_EMU(pil.height))
    marker = AnchorMarker(col=col-1, colOff=offset, row=row-1, rowOff=offset)
    xl.anchor = OneCellAnchor(_from=marker, ext=size)
    ws.add_image(xl)

# ── Sabit sütunlar (mevcut satış raporu — değişmedi) ─────────────────────────
COLUMNS = [
    ("Resim",      None,          8),
    ("Urun ID",    "prdID",      14),
    ("Kalem ID",   "itmID",      14),
    ("Etiket",     "tagPrice",   14),
    ("Satis",      "salePrice",  14),
    ("Indirim %",  "discount",   12),
    ("Tarih",      "date",       14),
    ("Lokasyon",   "location",   16),
]
TOTAL_KEYS = ["tagPrice", "salePrice"]

def build_workbook(rows):
    wb = Workbook()
    ws = wb.active
    ws.title = "Satis Raporu"

    for col_no, (label, _, width) in enumerate(COLUMNS, start=1):
        ws.column_dimensions[get_column_letter(col_no)].width = width
        cell = ws.cell(row=1, column=col_no, value=label)
        header_style(cell)
    ws.row_dimensions[1].height = 22

    totals = {k: 0.0 for k in TOTAL_KEYS}

    for idx, item in enumerate(rows):
        row_no = idx + 2
        ws.row_dimensions[row_no].height = ROW_PT

        for col_no, (_, key, _) in enumerate(COLUMNS, start=1):
            if key is None:
                b64 = item.get("image_base64", "")
                if b64:
                    try:
                        place_image(ws, b64, col_no, row_no, IMG_PX)
                    except Exception:
                        ws.cell(row=row_no, column=col_no, value="-")
                else:
                    ws.cell(row=row_no, column=col_no, value="-")
            else:
                val = item.get(key, "")
                cell = ws.cell(row=row_no, column=col_no, value=val)
                row_style(cell, idx)
                if key in totals:
                    try:
                        totals[key] += float(val)
                    except (ValueError, TypeError):
                        pass

    total_row = len(rows) + 2
    ws.row_dimensions[total_row].height = 24
    indirim = totals.get("tagPrice", 0) - totals.get("salePrice", 0)

    for col_no, (_, key, _) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=total_row, column=col_no)
        if col_no == 1:
            cell.value = "TOPLAM"
            cell.font      = Font(name="Arial", bold=True, size=10, color="FFFFFF")
            cell.fill      = PatternFill("solid", start_color=DARK)
            cell.alignment = Alignment(horizontal="center", vertical="center")
        elif key == "tagPrice":
            cell.value = totals["tagPrice"]
            cell.number_format = '#,##0.00'
            total_style(cell)
        elif key == "salePrice":
            cell.value = totals["salePrice"]
            cell.number_format = '#,##0.00'
            total_style(cell)
        elif key == "discount":
            cell.value = indirim
            cell.number_format = '#,##0.00'
            cell.font  = Font(name="Arial", bold=True, size=10, color="C00000")
            cell.fill  = PatternFill("solid", start_color="E8ECF4")
            s = Side(style="medium", color=DARK)
            cell.border = Border(left=s, right=s, top=s, bottom=s)
            cell.alignment = Alignment(horizontal="center", vertical="center")
        else:
            total_style(cell)

    ws.freeze_panes = "A2"
    last_col = get_column_letter(len(COLUMNS))
    ws.auto_filter.ref = f"A1:{last_col}1"
    return wb

# ── Dinamik workbook (yeni raporlar için) ────────────────────────────────────
def build_dynamic_workbook(rows, columns, title="Rapor"):
    """
    columns listesi şu formatta gelir:
    [
      {"label": "Resim",   "key": "image",   "width": 8,  "type": "image"},
      {"label": "Urun ID", "key": "prdID",   "width": 14, "type": "text"},
      {"label": "Tutar",   "key": "amount",  "width": 12, "type": "number", "total": true}
    ]
    type: "image" | "text" | "number"
    total: true olan number sütunları toplanır
    """
    wb = Workbook()
    ws = wb.active
    ws.title = title[:31]

    for col_no, col in enumerate(columns, start=1):
        ws.column_dimensions[get_column_letter(col_no)].width = col.get("width", 14)
        cell = ws.cell(row=1, column=col_no, value=col["label"])
        header_style(cell)
    ws.row_dimensions[1].height = 22

    total_keys = [c["key"] for c in columns
                  if c.get("total") and c.get("type") == "number"]
    totals = {k: 0.0 for k in total_keys}

    for idx, item in enumerate(rows):
        row_no = idx + 2
        ws.row_dimensions[row_no].height = ROW_PT

        for col_no, col in enumerate(columns, start=1):
            col_type = col.get("type", "text")
            key      = col.get("key", "")

            if col_type == "image":
                b64 = item.get("image_base64", "") or item.get(key, "")
                if b64:
                    try:
                        place_image(ws, b64, col_no, row_no, IMG_PX)
                    except Exception:
                        ws.cell(row=row_no, column=col_no, value="-")
                else:
                    ws.cell(row=row_no, column=col_no, value="-")
            elif col_type == "number":
                val = item.get(key, "")
                try:
                    val = float(val)
                    if key in totals:
                        totals[key] += val
                except (ValueError, TypeError):
                    pass
                cell = ws.cell(row=row_no, column=col_no, value=val)
                row_style(cell, idx)
            else:
                cell = ws.cell(row=row_no, column=col_no, value=item.get(key, ""))
                row_style(cell, idx)

    if total_keys:
        total_row = len(rows) + 2
        ws.row_dimensions[total_row].height = 24
        for col_no, col in enumerate(columns, start=1):
            cell = ws.cell(row=total_row, column=col_no)
            key  = col.get("key", "")
            if col_no == 1:
                cell.value = "TOPLAM"
                cell.font      = Font(name="Arial", bold=True, size=10, color="FFFFFF")
                cell.fill      = PatternFill("solid", start_color=DARK)
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif key in totals:
                cell.value = totals[key]
                cell.number_format = '#,##0.00'
                total_style(cell)
            else:
                total_style(cell)

    ws.freeze_panes = "A2"
    last_col = get_column_letter(len(columns))
    ws.auto_filter.ref = f"A1:{last_col}1"
    return wb

def parse_rows(req):
    body = req.get_json(force=True, silent=True)
    if not body or "data" not in body:
        return None, None
    rows = [r for r in body["data"]
            if any(str(v).strip() for k, v in r.items() if k != "image_base64")]
    return body, rows

def parse_dynamic_request(req):
    body = req.get_json(force=True, silent=True)
    if not body or "data" not in body or "columns" not in body:
        return None, None, None
    rows = [r for r in body["data"]
            if any(str(v).strip() for k, v in r.items() if k != "image_base64")]
    return body, body["columns"], rows

def save_to_tmp(wb, title="rapor"):
    token    = uuid.uuid4().hex[:6]
    tarih    = datetime.now().strftime("%Y-%m-%d")
    safe     = "".join(c if c.isalnum() else "_" for c in title.lower())[:20]
    filename = f"{safe}_{tarih}_{token}.xlsx"
    filepath = os.path.join(TMP_DIR, filename)
    wb.save(filepath)
    with _store_lock:
        _file_store[token] = {
            "path":       filepath,
            "filename":   filename,
            "created_at": time.time()
        }
    return token

# ── Mevcut endpoint: FM Pro, sabit sütunlar ──────────────────────────────────
@app.route("/generate-excel", methods=["POST"])
def generate_excel():
    ok, err = check_api_key()
    if not ok:
        return jsonify({"error": err}), 401
    body, rows = parse_rows(request)
    if body is None:
        return jsonify({"error": "JSON body eksik"}), 400
    if not rows:
        return jsonify({"error": "data listesi bos"}), 400
    wb  = build_workbook(rows)
    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return jsonify({"status": "ok", "rows": len(rows),
                    "excel_b64": base64.b64encode(out.read()).decode("utf-8")})

# ── Mevcut endpoint: WebDirect, sabit sütunlar ───────────────────────────────
@app.route("/generate-excel-file", methods=["POST"])
def generate_excel_file():
    ok, err = check_api_key()
    if not ok:
        return jsonify({"error": err}), 401
    _cleanup_old_files()
    body, rows = parse_rows(request)
    if body is None:
        return jsonify({"error": "JSON body eksik"}), 400
    if not rows:
        return jsonify({"error": "data listesi bos"}), 400
    wb    = build_workbook(rows)
    token = save_to_tmp(wb, "satis_raporu")
    base_url = request.host_url.rstrip("/")
    return jsonify({"status": "ok", "rows": len(rows),
                    "download_url": f"{base_url}/download/{token}"})

# ── Yeni endpoint: FM Pro, dinamik sütunlar ──────────────────────────────────
@app.route("/generate-excel-dynamic", methods=["POST"])
def generate_excel_dynamic():
    ok, err = check_api_key()
    if not ok:
        return jsonify({"error": err}), 401
    body, columns, rows = parse_dynamic_request(request)
    if body is None:
        return jsonify({"error": "JSON body eksik veya 'columns'/'data' yok"}), 400
    if not rows:
        return jsonify({"error": "data listesi bos"}), 400
    title = body.get("title", "Rapor")
    wb    = build_dynamic_workbook(rows, columns, title)
    out   = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return jsonify({"status": "ok", "rows": len(rows),
                    "excel_b64": base64.b64encode(out.read()).decode("utf-8")})

# ── Yeni endpoint: WebDirect, dinamik sütunlar ───────────────────────────────
@app.route("/generate-excel-dynamic-file", methods=["POST"])
def generate_excel_dynamic_file():
    ok, err = check_api_key()
    if not ok:
        return jsonify({"error": err}), 401
    _cleanup_old_files()
    body, columns, rows = parse_dynamic_request(request)
    if body is None:
        return jsonify({"error": "JSON body eksik veya 'columns'/'data' yok"}), 400
    if not rows:
        return jsonify({"error": "data listesi bos"}), 400
    title = body.get("title", "Rapor")
    wb    = build_dynamic_workbook(rows, columns, title)
    token = save_to_tmp(wb, title)
    base_url = request.host_url.rstrip("/")
    return jsonify({"status": "ok", "rows": len(rows),
                    "download_url": f"{base_url}/download/{token}"})

# ── Download endpoint ────────────────────────────────────────────────────────
@app.route("/download/<token>", methods=["GET"])
def download_file(token):
    with _store_lock:
        entry = _file_store.get(token)
    if not entry or not os.path.exists(entry["path"]):
        return jsonify({"error": "Dosya bulunamadi veya suresi doldu"}), 404
    return send_file(entry["path"], as_attachment=True,
                     download_name=entry["filename"],
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

# ── Health ───────────────────────────────────────────────────────────────────
@app.route("/health", methods=["GET"])
def health():
    return jsonify({"status": "ok"})

if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5000))
    app.run(host="0.0.0.0", port=port)
